from openpyxl import load_workbook
import os
from conftest import RESOURCE_PATH
# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def xlsx_test():
    workbook = load_workbook(os.path.join(RESOURCE_PATH, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'